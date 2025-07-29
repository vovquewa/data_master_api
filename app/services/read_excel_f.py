import os
from typing import List, Optional, Tuple
import pandas as pd 
from openpyxl import load_workbook
from xlrd import open_workbook
from openpyxl.cell import MergedCell


class FileFinder:
    def __init__(self, directory: str):
        """
        Инициализирует объект FileFinder с указанной директорией.
        
        :param directory: Путь к директории для поиска файлов
        """
        self.directory = directory
        
    def find_files_by_partial_name(self, partial_name: str, case_sensitive: bool = False) -> List[str]:
        """
        Находит все файлы в директории, содержащие указанную часть названия.
        
        :param partial_name: Часть названия файла для поиска
        :param case_sensitive: Учитывать регистр при поиске (по умолчанию False)
        :return: Список полных путей к найденным файлам
        """
        if not os.path.isdir(self.directory):
            raise ValueError(f"Указанная директория не существует: {self.directory}")
            
        found_files = []
        partial_name = partial_name if case_sensitive else partial_name.lower()
        
        for filename in os.listdir(self.directory):
            filepath = os.path.join(self.directory, filename)
            if os.path.isfile(filepath):
                compare_name = filename if case_sensitive else filename.lower()
                if partial_name in compare_name:
                    found_files.append(filepath)
                    
        return found_files
    
    def find_files_by_extension(self, extension: str) -> List[str]:
        """
        Находит все файлы с указанным расширением.
        
        :param extension: Расширение файла (например, 'txt', 'csv')
        :return: Список полных путей к найденным файлам
        """
        if not extension.startswith('.'):
            extension = '.' + extension
            
        return self.find_files_by_partial_name(extension)

class HeaderFinder:
    def __init__(self, file_path: str, search_terms: List[str]):
        """
        Универсальный поиск заголовков в Excel файлах (XLS/XLSX).
        
        :param file_path: Путь к Excel файлу
        :param search_terms: Список терминов для поиска в заголовках
        """
        if not os.path.exists(file_path):
            raise FileNotFoundError(f"Файл не найден: {file_path}")
            
        self.file_path = file_path
        self.search_terms = [term.lower() for term in search_terms]
        self.header_row = None
        self.file_type = self._detect_file_type()

    def _detect_file_type(self) -> str:
        """Определяет тип файла по расширению"""
        ext = os.path.splitext(self.file_path)[1].lower()
        if ext == '.xlsx':
            return 'xlsx'
        elif ext == '.xls':
            return 'xls'
        else:
            raise ValueError(f"Неподдерживаемый формат файла: {ext}")

    def find_header(self, max_rows_to_check: int = 30) -> Optional[int]:
        """
        Находит номер строки с заголовками.
        
        :param max_rows_to_check: Максимальное количество строк для проверки
        :return: Номер строки (начиная с 1) или None
        """
        if self.file_type == 'xlsx':
            return self._find_header_xlsx(max_rows_to_check)
        else:
            return self._find_header_xls(max_rows_to_check)

    def _find_header_xlsx(self, max_rows_to_check: int) -> Optional[int]:
        """Поиск заголовков в XLSX файле"""
        wb = load_workbook(self.file_path, read_only=True)
        ws = wb.active
        
        for row_idx, row in enumerate(ws.iter_rows(max_row=max_rows_to_check), 1):
            cell_values = [str(cell.value).lower() if cell.value else "" for cell in row]
            if self._row_contains_terms(cell_values):
                self.header_row = row_idx
                wb.close()
                return row_idx
                
        wb.close()
        return None

    def _find_header_xls(self, max_rows_to_check: int) -> Optional[int]:
        """Поиск заголовков в XLS файле"""
        wb = open_workbook(self.file_path)
        ws = wb.sheet_by_index(0)
        
        for row_idx in range(min(max_rows_to_check, ws.nrows)):
            row = ws.row(row_idx)
            cell_values = [str(cell.value).lower() if cell.value else "" for cell in row]
            if self._row_contains_terms(cell_values):
                self.header_row = row_idx + 1  # xlrd использует 0-based индексы
                return self.header_row
                
        return None

    def _row_contains_terms(self, cell_values: List[str]) -> bool:
        """Проверяет, содержит ли строка искомые термины"""
        matches = sum(1 for term in self.search_terms 
                     if any(term in cell_value for cell_value in cell_values))
        return matches >= len(self.search_terms) // 2  # Хотя бы половина терминов

    def get_data(self) -> Tuple[Optional[int], List[dict]]:
        """
        Возвращает данные из файла.
        
        :return: Кортеж (номер строки с заголовками, список словарей с данными)
        """
        if self.header_row is None:
            self.find_header()
            
        if self.header_row is None:
            return None, []
            
        if self.file_type == 'xlsx':
            return self._get_data_xlsx()
        else:
            return self._get_data_xls()

    def _get_data_xlsx(self) -> Tuple[int, List[dict]]:
        """Получение данных из XLSX файла"""
        wb = load_workbook(self.file_path)
        ws = wb.active
        
        # Получаем заголовки
        headers = []
        for cell in ws[self.header_row]:
            self.header_row
            if isinstance(cell, MergedCell):
                unique_values = set()
                for merged_range in ws.merged_cells.ranges:
                 # Берём значение только из первой ячейки объединённого блока
                    first_cell = ws.cell(self.header_row, merged_range.min_col)
                   # last_cell = ws.cell(self.header_row, merged_range.max_col)
                  
                    if first_cell.value:
                        unique_values.add(str(first_cell.value))

                # Объединяем уникальные значения
                merged_result = "; ".join(unique_values).strip()
              
                
                cell = ws.merged_cells.ranges.clear()
                if cell != None:  
                    headers =headers.append(str(merged_result) if merged_result!=None else f"column_{cell.column_letter}")   
            else:
                headers.append(str(cell.value) if cell.value else f"column_{cell.column_letter}")
        
        # Собираем данные
        data = []
        for row in ws.iter_rows(min_row=self.header_row + 1):
            row_data = {}
            for i, cell in enumerate(row):
                if i < len(headers):
                    row_data[headers[i]] = cell.value
            data.append(row_data)
            
        wb.close()
        return self.header_row, data

    def _get_data_xls(self) -> Tuple[int, List[dict]]:
        """Получение данных из XLS файла"""
        wb = open_workbook(self.file_path)
        ws = wb.sheet_by_index(0)
        
        # Получаем заголовки (xlrd использует 0-based индексы)
        header_row_idx = self.header_row - 1
        headers = []
        for i in range(ws.ncols):
            cell_value = ws.cell_value(header_row_idx, i)
            headers.append(str(cell_value) if cell_value else f"column_{i+1}")
        
        # Собираем данные
        data = []
        for row_idx in range(header_row_idx + 1, ws.nrows):
            row_data = {}
            for col_idx in range(ws.ncols):
                if col_idx < len(headers):
                    row_data[headers[col_idx]] = ws.cell_value(row_idx, col_idx)
            data.append(row_data)
            
        return self.header_row, data

    def to_dataframe(self):
        """Конвертирует данные в pandas DataFrame (требуется установленный pandas)"""
        try:
            import pandas as pd
        except ImportError:
            raise ImportError("Для использования этого метода необходимо установить pandas")
        dtyp = ''    
        row_num, data = self.get_data()
        return pd.DataFrame(data)
        
#home_dir = os.getcwd()   
#dir_rep = FileFinder('/'.join([home_dir, 'Общее']))
#supppliers_files = dir_rep.find_files_by_partial_name('код')[0]
#order_files = dir_rep.find_files_by_partial_name('заказ')[0]
#dfor = pd.read_excel(order_files)
#order_head = HeaderFinder(supppliers_files, ['Номенклатура', 'BOOK_ID',	'КИЗ'])

#clo = order_head.to_dataframe()[['Номенклатура', 'BOOK_ID',	'КИЗ']]
#print(clo)